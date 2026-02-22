# mrp_routes.py
# Rotas para o sistema MRP (Material Requirements Planning)

from flask import render_template, request, jsonify, send_file, session, flash, redirect, url_for
import psycopg2
from psycopg2 import Error
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from config import DB_HOST, DB_NAME, DB_USER, DB_PASS, DB_PORT, SIGA_DB_NAME, MRP_DB_NAME


def get_db_connection(db_name=MRP_DB_NAME):
    """Cria e retorna uma conexão com o banco de dados PostgreSQL."""
    try:
        connection = psycopg2.connect(
            host=DB_HOST,
            database=db_name,
            user=DB_USER,
            password=DB_PASS,
            port=DB_PORT
        )
        return connection
    except Error as e:
        print(f"Erro ao conectar ao banco de dados {db_name}: {e}")
        return None


def get_estoques_bulk(cursor, produtos_codes):
    """Busca estoques calculando: último saldo possali1 + movimentações posteriores da toqmovi."""
    if not produtos_codes:
        return {}

    try:
        placeholders = ','.join(['%s'] * len(produtos_codes))

        # Query otimizada: calcula tudo de uma vez com CTEs
        query = f"""
            WITH ultimos_saldos AS (
                SELECT DISTINCT ON (produto) produto, psisaldo, psidata
                FROM possali1
                WHERE produto IN ({placeholders})
                ORDER BY produto, psidata DESC
            ),
            movimentacoes_posteriores AS (
                SELECT
                    t.priproduto AS produto,
                    COALESCE(SUM(
                        CASE
                            WHEN t.pritransac IN (3, 6) THEN t.priquanti
                            WHEN t.pritransac IN (14, 16) THEN -t.priquanti
                            ELSE 0
                        END
                    ), 0) AS saldo_movimentacoes
                FROM toqmovi t
                INNER JOIN ultimos_saldos us ON t.priproduto = us.produto
                WHERE t.pridata > us.psidata
                GROUP BY t.priproduto
            )
            SELECT
                us.produto,
                COALESCE(us.psisaldo, 0) + COALESCE(mp.saldo_movimentacoes, 0) AS estoque_final
            FROM ultimos_saldos us
            LEFT JOIN movimentacoes_posteriores mp ON us.produto = mp.produto
        """

        cursor.execute(query, produtos_codes)
        results = cursor.fetchall()
        estoques_dict = {row[0]: float(row[1] or 0) for row in results}

        print(f"DEBUG ESTOQUES: Calculados estoques para {len(estoques_dict)} produtos (saldo base + movimentações)")
        return estoques_dict
    except Error as e:
        print(f"Erro ao buscar estoques: {e}")
        import traceback
        traceback.print_exc()
        return {}


def get_ordens_bulk(cursor, produtos_codes, lotes_producao_selecionados=None):
    """Busca ordens PENDENTES de múltiplos produtos SEMPRE de TODAS as OPs OSSO em aberto.

    Calcula: Ordem Total - Produção já Realizada = Ordem Pendente

    OPs OSSO são ordens de fabricação (entrada), por isso sempre consideramos todas.

    Args:
        cursor: Cursor do banco de dados
        produtos_codes: Lista de códigos de produtos
        lotes_producao_selecionados: Não usado, mantido por compatibilidade
    """
    if not produtos_codes:
        return {}

    try:
        # SEMPRE buscar TODAS as OPs OSSO em aberto (não depende de seleção do usuário)
        print(f"DEBUG ORDENS: Buscando ordens de TODAS as OPs OSSO em aberto...")

        # Buscar todos os lotes OSSO em aberto
        query_lotes = """
            SELECT lotcod
            FROM loteprod
            WHERE lotstatus IN ('EP', 'ES')
              AND lotdes ILIKE %s
              AND lotdes ILIKE %s
        """
        cursor.execute(query_lotes, ('%OSSO%', '%OP%'))
        lotes_osso = [row[0] for row in cursor.fetchall()]

        if not lotes_osso or len(lotes_osso) == 0:
            print(f"DEBUG ORDENS: Nenhum lote OSSO encontrado no sistema, retornando 0 ordens")
            return {}

        print(f"DEBUG ORDENS: Encontrados {len(lotes_osso)} lotes OSSO no sistema")

        placeholders_produtos = ','.join(['%s'] * len(produtos_codes))
        placeholders_lotes = ','.join(['%s'] * len(lotes_osso))

        # Query otimizada: calcula ordem total - produção já realizada
        query = f"""
            SELECT
                o.ordproduto,
                SUM(o.ordquanti - COALESCE(t.qtd_produzida, 0)) AS ordem_pendente
            FROM ordem o
            LEFT JOIN (
                SELECT priordem, priproduto, SUM(priquanti) AS qtd_produzida
                FROM toqmovi
                WHERE pritransac IN (3, 6)  -- Apenas entradas (produção)
                GROUP BY priordem, priproduto
            ) t ON o.ordem = t.priordem AND o.ordproduto = t.priproduto
            WHERE o.ordproduto IN ({placeholders_produtos})
              AND o.lotcod IN ({placeholders_lotes})
              AND (o.orddtence IS NULL OR o.orddtence = '0001-01-01')
            GROUP BY o.ordproduto
            HAVING SUM(o.ordquanti - COALESCE(t.qtd_produzida, 0)) > 0
        """
        params_query = tuple(produtos_codes) + tuple(lotes_osso)
        cursor.execute(query, params_query)

        results = cursor.fetchall()
        ordens_dict = {row[0]: float(row[1] or 0) for row in results}
        print(f"DEBUG ORDENS: Encontradas ordens PENDENTES para {len(ordens_dict)} produtos")
        if len(ordens_dict) > 0:
            sample = list(ordens_dict.items())[:3]
            print(f"DEBUG ORDENS: Amostra (ordem pendente): {sample}")
        return ordens_dict
    except Error as e:
        print(f"Erro ao buscar ordens: {e}")
        import traceback
        traceback.print_exc()
        return {}


def get_reservas_bulk(cursor, produtos_codes, lotes_producao_selecionados=None):
    """Busca reservas PENDENTES de múltiplos produtos de uma vez (bulk query).

    Calcula: Requisitado - Movimentado = Reserva Pendente

    Args:
        cursor: Cursor do banco de dados
        produtos_codes: Lista de códigos de produtos
        lotes_producao_selecionados: Lista de códigos de lotes de produção selecionados pelo usuário
    """
    if not produtos_codes:
        return {}

    try:
        # Filtrar lotes pela DESCRIÇÃO (lotdes) na tabela loteprod
        if lotes_producao_selecionados and len(lotes_producao_selecionados) > 0:
            placeholders_lotes = ','.join(['%s'] * len(lotes_producao_selecionados))

            # Buscar lotes cuja descrição contém 'OP' e ('PETRA' ou 'GARLAND' ou 'SOLARE')
            query_lotes = f"""
                SELECT lotcod
                FROM loteprod
                WHERE lotcod IN ({placeholders_lotes})
                  AND lotdes ILIKE %s
                  AND (lotdes ILIKE %s OR lotdes ILIKE %s OR lotdes ILIKE %s)
            """
            params_lotes = tuple(lotes_producao_selecionados) + ('%OP%', '%PETRA%', '%GARLAND%', '%SOLARE%')
            cursor.execute(query_lotes, params_lotes)
            lotes_saida = [row[0] for row in cursor.fetchall()]

            if not lotes_saida or len(lotes_saida) == 0:
                print(f"DEBUG RESERVAS: Nenhum lote PETRA/GARLAND/SOLARE encontrado entre os selecionados, retornando 0 reservas")
                print(f"DEBUG RESERVAS: Lotes selecionados: {lotes_producao_selecionados}")
                return {}

            print(f"DEBUG RESERVAS: Encontrados {len(lotes_saida)} lotes de saída: {lotes_saida[:3]}...")

            placeholders_produtos = ','.join(['%s'] * len(produtos_codes))
            placeholders_lotes = ','.join(['%s'] * len(lotes_saida))

            # Query otimizada: primeiro busca requisições, depois movimentações apenas das requisições encontradas
            # PASSO 1: Buscar requisições dos lotes selecionados
            query_req = f"""
                SELECT r.reqord, r.reqproduto, r.rqoquanti
                FROM reqordem r
                INNER JOIN ordem o ON r.reqord = o.ordem
                WHERE r.reqproduto IN ({placeholders_produtos})
                  AND o.lotcod IN ({placeholders_lotes})
                  AND (o.orddtence IS NULL OR o.orddtence = '0001-01-01')
            """
            params_query = tuple(produtos_codes) + tuple(lotes_saida)
            cursor.execute(query_req, params_query)
            requisicoes = cursor.fetchall()

            if not requisicoes:
                print(f"DEBUG RESERVAS: Nenhuma requisição encontrada nos lotes selecionados")
                return {}

            # PASSO 2: Buscar movimentações apenas das ordens requisitadas
            ordens_requisitadas = list(set([req[0] for req in requisicoes]))

            if len(ordens_requisitadas) > 0:
                placeholders_ordens = ','.join(['%s'] * len(ordens_requisitadas))
                query_mov = f"""
                    SELECT priordem, priproduto, SUM(priquanti) AS qtd_movimentada
                    FROM toqmovi
                    WHERE priordem IN ({placeholders_ordens})
                    GROUP BY priordem, priproduto
                """
                cursor.execute(query_mov, tuple(ordens_requisitadas))
                movimentacoes = {(row[0], row[1]): float(row[2]) for row in cursor.fetchall()}
            else:
                movimentacoes = {}

            # PASSO 3: Calcular pendentes em Python (convertendo para float)
            reservas_dict = {}
            for reqord, reqproduto, rqoquanti in requisicoes:
                qtd_movimentada = movimentacoes.get((reqord, reqproduto), 0)
                pendente = float(rqoquanti) - float(qtd_movimentada)
                if pendente > 0:
                    reservas_dict[reqproduto] = reservas_dict.get(reqproduto, 0) + pendente
        else:
            # Se não houver lotes selecionados, não calcular reservas
            # Reservas só são relevantes para lotes PETRA/GARLAND/SOLARE selecionados
            print(f"DEBUG RESERVAS: Nenhum lote de produção selecionado, retornando 0 reservas")
            reservas_dict = {}

        # Log de debug
        print(f"DEBUG RESERVAS: Buscados {len(produtos_codes)} produtos")
        print(f"DEBUG RESERVAS: Encontradas reservas PENDENTES para {len(reservas_dict)} produtos")
        if len(reservas_dict) > 0:
            # Mostrar primeiros 5 produtos com reserva
            sample = list(reservas_dict.items())[:5]
            print(f"DEBUG RESERVAS: Amostra: {sample}")

        return reservas_dict
    except Error as e:
        print(f"Erro ao buscar reservas: {e}")
        import traceback
        traceback.print_exc()
        return {}


def get_mrp_filters():
    """Busca os filtros disponíveis para o cálculo MRP."""
    connection = get_db_connection()
    if not connection:
        return {'pedidos': [], 'lotes_carga': [], 'lotes_producao': []}

    pedidos = []
    lotes_carga = []
    lotes_producao = []

    try:
        cursor = connection.cursor()

        # Buscar pedidos de venda em aberto
        try:
            cursor.execute("""
                SELECT DISTINCT pedido, pedcliente, peddata
                FROM pedido
                WHERE pedaprova = 'S'
                  AND (pedsitua IS NULL OR pedsitua = '' OR pedsitua = 'P')
                ORDER BY peddata DESC
                LIMIT 100
            """)
            pedidos = [{'codigo': row[0], 'cliente': row[1], 'data': row[2]} for row in cursor.fetchall()]
        except Error as e:
            print(f"Aviso: Tabela 'pedido' não encontrada ou erro ao buscar pedidos: {e}")
            pedidos = []

        # Buscar lotes de carga (apenas PETRA/GARLAND/SOLARE)
        try:
            cursor.execute("""
                SELECT DISTINCT lcacod, lcades, lcaprev
                FROM lotecar
                WHERE lcades ILIKE '%PETRA%'
                   OR lcades ILIKE '%GARLAND%'
                   OR lcades ILIKE '%SOLARE%'
                ORDER BY lcaprev DESC
                LIMIT 100
            """)
            lotes_carga = [{'codigo': row[0], 'descricao': row[1], 'previsao': row[2]} for row in cursor.fetchall()]
        except Error as e:
            print(f"Aviso: Tabela 'lotecar' não encontrada ou erro ao buscar lotes de carga: {e}")
            lotes_carga = []

        # Buscar lotes de produção em aberto (excluindo AVULSO e OSSO)
        # OPs OSSO são usadas automaticamente para ordens, não precisam ser selecionadas
        try:
            cursor.execute("""
                SELECT DISTINCT lotcod, lotdes, lotdtini
                FROM loteprod
                WHERE lotstatus IN ('EP', 'ES')
                  AND (lotdes NOT ILIKE '%AVULSO%' OR lotdes IS NULL)
                  AND lotdes NOT ILIKE '%OSSO%'
                ORDER BY lotdtini DESC
                LIMIT 100
            """)
            lotes_producao = [{'codigo': row[0], 'descricao': row[1], 'data': row[2]} for row in cursor.fetchall()]
        except Error as e:
            print(f"Aviso: Tabela 'loteprod' não encontrada ou erro ao buscar lotes de produção: {e}")
            lotes_producao = []

        cursor.close()
        connection.close()

        return {
            'pedidos': pedidos,
            'lotes_carga': lotes_carga,
            'lotes_producao': lotes_producao
        }
    except Error as e:
        print(f"Erro ao buscar filtros MRP: {e}")
        if connection:
            connection.close()
        return {
            'pedidos': pedidos,
            'lotes_carga': lotes_carga,
            'lotes_producao': lotes_producao
        }


def get_cubagem_madeira_bulk(cursor, produtos_codes):
    """
    Calcula a cubagem (M³) de madeira para múltiplos produtos.

    Tabela: dimensao
    - dimcodigo = código do produto
    - dimquanti = quantidade de peças
    - dimbrucomp = comprimento bruto (mm)
    - dimbrularg = largura bruta (mm)
    - dimbruespe = espessura bruta (mm)

    Fórmula: M³ = (dimbrucomp/1000) * (dimbrularg/1000) * (dimbruespe/1000) * dimquanti

    Returns:
        Dict: {produto_code: cubagem_m3}
    """
    if not produtos_codes:
        return {}

    try:
        placeholders = ','.join(['%s'] * len(produtos_codes))
        query = f"""
            SELECT
                TRIM(dimcodigo) as produto,
                SUM(
                    (COALESCE(dimbrucomp, 0) / 1000.0) *
                    (COALESCE(dimbrularg, 0) / 1000.0) *
                    (COALESCE(dimbruespe, 0) / 1000.0) *
                    COALESCE(dimquanti, 0)
                ) as cubagem_m3
            FROM dimensao
            WHERE TRIM(dimcodigo) IN ({placeholders})
            GROUP BY TRIM(dimcodigo)
        """
        cursor.execute(query, tuple([p.strip() for p in produtos_codes]))
        results = cursor.fetchall()

        cubagem_dict = {row[0]: float(row[1] or 0) for row in results}
        print(f"DEBUG CUBAGEM: Calculada cubagem para {len(cubagem_dict)} produtos")
        return cubagem_dict
    except Error as e:
        print(f"Erro ao buscar cubagem: {e}")
        import traceback
        traceback.print_exc()
        return {}


def get_material_types_bulk(cursor, produtos_codes):
    """
    Determina o tipo de matéria-prima de múltiplos produtos de uma vez baseado na estrutura.

    Tabela: estrutur
    - estproduto = item pai (produto principal)
    - estfilho = item filho (componente)

    Tabela: produto
    - grupo = grupo do produto
    - subgrupo = subgrupo do produto

    Regras:
    - Grupo 10, Subgrupo 2 = Compensado
    - Grupo 10, Subgrupo 3 = Madeira
    - Grupo 10, Subgrupo 1 = MDF

    Returns:
        Dict: {produto_code: 'compensado'/'madeira'/'mdf'}
    """
    if not produtos_codes:
        return {}

    try:
        placeholders = ','.join(['%s'] * len(produtos_codes))
        query = f"""
            SELECT DISTINCT
                TRIM(e.estproduto) as produto_pai,
                p.subgrupo as subgrupo_filho
            FROM estrutur e
            INNER JOIN produto p ON TRIM(e.estfilho) = TRIM(p.produto)
            WHERE TRIM(e.estproduto) IN ({placeholders})
              AND p.grupo = 10
              AND p.subgrupo IN (1, 2, 3)
        """
        cursor.execute(query, tuple([p.strip() for p in produtos_codes]))
        results = cursor.fetchall()

        print(f"DEBUG MATERIAL: Query retornou {len(results)} registros")
        if len(results) > 0:
            print(f"DEBUG MATERIAL: Primeiros 5 resultados: {results[:5]}")

        # Mapear produtos para tipos de material
        material_types = {}
        for row in results:
            produto_code = row[0].strip() if row[0] else ''
            subgrupo = int(row[1]) if row[1] else 0

            # Só adicionar se ainda não foi mapeado (prioridade: primeiro encontrado)
            if produto_code and produto_code not in material_types:
                if subgrupo == 2:
                    material_types[produto_code] = 'compensado'
                elif subgrupo == 3:
                    material_types[produto_code] = 'madeira'
                elif subgrupo == 1:
                    material_types[produto_code] = 'mdf'

        print(f"DEBUG MATERIAL: Encontrados {len(material_types)} produtos com matéria-prima definida")
        if len(material_types) > 0:
            sample = list(material_types.items())[:5]
            print(f"DEBUG MATERIAL: Amostra mapeada: {sample}")
        return material_types
    except Exception as e:
        print(f"Erro ao buscar tipos de material: {e}")
        import traceback
        traceback.print_exc()
        return {}


def calculate_mrp(filters):
    """
    Calcula o MRP baseado nos filtros selecionados.

    Args:
        filters: Dict contendo os filtros aplicados
            - pedidos: Lista de códigos de pedidos
            - lotes_carga: Lista de códigos de lotes de carga
            - lotes_producao: Lista de códigos de lotes de produção
            - material_filter: Filtro específico (compensado, madeira, mdf)

    Returns:
        Lista de produtos com cálculo MRP
    """
    connection = get_db_connection()
    if not connection:
        return []

    try:
        cursor = connection.cursor()

        # Definir timeout de 5 minutos (view pode demorar na primeira execução)
        cursor.execute("SET statement_timeout = '300s'")

        # Determinar o nome da view baseado no banco de dados
        # madresilva usa madres_vw_mrp_base, outros bancos usam vw_mrp_base
        view_name = "madres_vw_mrp_base" if MRP_DB_NAME == "madresilva" else "vw_mrp_base"
        print(f"Usando view: {view_name} do banco {MRP_DB_NAME}")

        # Buscar todos os produtos da view base
        query = f"""
            SELECT
                produto,
                descricao,
                grupo_abrev,
                estoque_minimo,
                lote_economico,
                saldo_estoque,
                ordens,
                reservas,
                proorigem
            FROM {view_name}
            WHERE 1=1
            ORDER BY grupo_abrev, produto
        """

        print(f"Executando query MRP...")
        cursor.execute(query)
        print("Query executada com sucesso")
        rows = cursor.fetchall()
        print(f"Total de produtos encontrados: {len(rows)}")

        # Buscar estoques, ordens e reservas em queries separadas (mais rápido)
        produtos_codes = [row[0] for row in rows]

        # Obter lotes de produção selecionados pelo usuário
        lotes_producao_selecionados = filters.get('lotes_producao', [])
        print(f"Lotes de produção selecionados: {lotes_producao_selecionados}")

        # Buscar estoques de todos os produtos de uma vez
        print("Buscando estoques...")
        estoques_dict = get_estoques_bulk(cursor, produtos_codes)

        # Buscar ordens de todos os produtos de uma vez (filtrado por lotes OSSO)
        print("Buscando ordens...")
        ordens_dict = get_ordens_bulk(cursor, produtos_codes, lotes_producao_selecionados)

        # Buscar reservas de todos os produtos de uma vez (filtrado por lotes PETRA/GARLAND/SOLARE)
        print("Buscando reservas...")
        reservas_dict = get_reservas_bulk(cursor, produtos_codes, lotes_producao_selecionados)

        # Buscar tipos de material de todos os produtos de uma vez
        print("Buscando tipos de material...")
        material_types = get_material_types_bulk(cursor, produtos_codes)

        # Buscar cubagem de madeira de todos os produtos de uma vez
        print("Buscando cubagem de madeira...")
        cubagens_dict = get_cubagem_madeira_bulk(cursor, produtos_codes)

        print("Processando dados...")
        produtos = []
        for row in rows:
            produto_code = row[0]
            descricao = row[1]
            grupo_abrev = row[2] or ''
            estoque_min = float(row[3] or 0)
            lote_eco = float(row[4] or 0)
            proorigem = row[8] or ''

            # Buscar valores nos dicionários
            estoque = estoques_dict.get(produto_code, 0)
            ordens = ordens_dict.get(produto_code, 0)
            reservas = reservas_dict.get(produto_code, 0)

            # Calcular categoria do produto (versão simplificada e rápida)
            categoria = determine_product_category_simple(produto_code, descricao, proorigem)

            # Cálculo MRP
            saldo = estoque + ordens - reservas

            # Determinar classificação e quantidade a produzir
            sem_necessidade = 0
            necessidade = 0
            urgente = 0

            if saldo < 0:
                # Urgente: saldo negativo, não consegue atender demanda
                # Fórmula: abs(saldo) + estoque_minimo
                # Não arredonda para lote econômico, usa valor exato
                quantidade_necessaria = abs(saldo) + estoque_min
                urgente = quantidade_necessaria
            elif saldo < estoque_min:
                # Necessidade: saldo positivo mas abaixo do estoque mínimo
                quantidade_necessaria = estoque_min - saldo
                # Ajustar ao lote econômico
                if lote_eco > 0 and quantidade_necessaria < lote_eco:
                    necessidade = lote_eco
                else:
                    # Se lote econômico for 0, usar quantidade necessária
                    necessidade = quantidade_necessaria
            else:
                # Sem necessidade: saldo atende e está acima do estoque mínimo
                sem_necessidade = 0

            # Saldo - Reserva: Estoque + Ordem - Reserva
            saldo_menos_reserva = estoque + ordens - reservas
            # Saldo Futuro: Estoque + Ordem - Reserva + OP (OP será adicionado no frontend)
            saldo_futuro_calc = estoque + ordens - reservas

            # Obter tipo de material do produto (remover espaços para fazer o match)
            material_type = material_types.get(produto_code.strip(), '')

            # Obter cubagem unitária do produto (M³ por peça)
            cubagem_unitaria = cubagens_dict.get(produto_code.strip(), 0)

            produto_dict = {
                'codigo': produto_code,
                'descricao': descricao,
                'grupo': grupo_abrev[:5],
                'estoque_minimo': estoque_min,
                'lote_economico': lote_eco,
                'estoque': estoque,
                'ordem': ordens,
                'reserva': reservas,
                'sem_necessidade': sem_necessidade,
                'necessidade': necessidade,
                'urgente': urgente,
                'op': 0,  # Será preenchido pelo usuário
                'saldo_menos_reserva': saldo_menos_reserva,
                'saldo_futuro': saldo_futuro_calc,
                'categoria': categoria,
                'material_type': material_type,  # Tipo de matéria prima: compensado, madeira, mdf
                'cubagem_unitaria': cubagem_unitaria  # M³ por peça
            }

            produtos.append(produto_dict)

        cursor.close()
        connection.close()

        return produtos

    except Error as e:
        print(f"Erro ao calcular MRP: {e}")
        if connection:
            connection.close()
        return []


def get_orders_reservations_by_lot(cursor, produto_code, lotes_producao):
    """
    Calcula ordens e reservas específicas baseado nos lotes de produção selecionados.

    Lotes OSSO (pela descrição): Entrada (+)
    Lotes PETRA, GARLAND, SOLARE (pela descrição): Saída (-)
    """
    ordens = 0
    reservas = 0

    try:
        # Ordens (entrada) - lotes cuja descrição contém OP e OSSO
        if lotes_producao and len(lotes_producao) > 0:
            placeholders_lotes = ','.join(['%s'] * len(lotes_producao))
            query_lotes_entrada = f"""
                SELECT lotcod
                FROM loteprod
                WHERE lotcod IN ({placeholders_lotes})
                  AND lotdes ILIKE %s
                  AND lotdes ILIKE %s
            """
            params_lotes = tuple(lotes_producao) + ('%OSSO%', '%OP%')
            cursor.execute(query_lotes_entrada, params_lotes)
            lotes_entrada = [row[0] for row in cursor.fetchall()]

            if lotes_entrada and len(lotes_entrada) > 0:
                placeholders = ','.join(['%s'] * len(lotes_entrada))
                query_ordens = f"""
                    SELECT COALESCE(SUM(ordquanti), 0)
                    FROM ordem
                    WHERE ordproduto = %s
                      AND lotcod IN ({placeholders})
                      AND (orddtence IS NULL OR orddtence = '0001-01-01')
                """
                params_ordens = tuple([produto_code]) + tuple(lotes_entrada)
                cursor.execute(query_ordens, params_ordens)
                result = cursor.fetchone()
                ordens = float(result[0] or 0)

        # Reservas (saída) - lotes cuja descrição contém OP e (PETRA ou GARLAND ou SOLARE)
        # Calcula apenas requisições PENDENTES (requisitado - movimentado)
        if lotes_producao and len(lotes_producao) > 0:
            placeholders_lotes = ','.join(['%s'] * len(lotes_producao))
            query_lotes_saida = f"""
                SELECT lotcod
                FROM loteprod
                WHERE lotcod IN ({placeholders_lotes})
                  AND lotdes ILIKE %s
                  AND (lotdes ILIKE %s OR lotdes ILIKE %s OR lotdes ILIKE %s)
            """
            params_lotes = tuple(lotes_producao) + ('%OP%', '%PETRA%', '%GARLAND%', '%SOLARE%')
            cursor.execute(query_lotes_saida, params_lotes)
            lotes_saida = [row[0] for row in cursor.fetchall()]

            if lotes_saida and len(lotes_saida) > 0:
                placeholders = ','.join(['%s'] * len(lotes_saida))
                query_reservas = f"""
                    SELECT COALESCE(SUM(r.rqoquanti - COALESCE(t.qtd_movimentada, 0)), 0)
                    FROM reqordem r
                    INNER JOIN ordem o ON r.reqord = o.ordem
                    LEFT JOIN (
                        SELECT priordem, priproduto, SUM(priquanti) AS qtd_movimentada
                        FROM toqmovi
                        GROUP BY priordem, priproduto
                    ) t ON r.reqord = t.priordem AND r.reqproduto = t.priproduto
                    WHERE r.reqproduto = %s
                      AND o.lotcod IN ({placeholders})
                      AND (o.orddtence IS NULL OR o.orddtence = '0001-01-01')
                      AND (r.rqoquanti - COALESCE(t.qtd_movimentada, 0)) > 0
                """
                params_reservas = tuple([produto_code]) + tuple(lotes_saida)
                cursor.execute(query_reservas, params_reservas)
                result = cursor.fetchone()
                reservas = float(result[0] or 0)

    except Error as e:
        print(f"Erro ao calcular ordens/reservas por lote: {e}")

    return ordens, reservas


def determine_product_category_simple(produto_code, descricao, proorigem):
    """
    Determina a categoria do produto baseado nas regras de negócio (versão simplificada).

    Categorias:
    - Tampos: código contém OSS.TP
    - Assento/Encosto: descrição começa com ASSENTO* ou ENCOSTO*
    - Montado: produto fabricado (proorigem = 'F')
    - Componentes: demais itens
    """
    descricao_upper = (descricao or '').upper()
    codigo_upper = (produto_code or '').upper()

    # Tampos
    if 'OSS.TP' in codigo_upper:
        return 'Tampos'

    # Assento/Encosto
    if descricao_upper.startswith('ASSENTO') or descricao_upper.startswith('ENCOSTO'):
        return 'Assento/Encosto'

    # Montado - produto fabricado
    if proorigem == 'F':
        return 'Montado'

    # Componentes (padrão)
    return 'Componentes'


def determine_product_category(cursor, produto_code, descricao):
    """
    DEPRECATED: Use determine_product_category_simple() para melhor performance.
    """
    # Por compatibilidade, usar a versão simples
    return determine_product_category_simple(produto_code, descricao, '')


def save_mrp_to_database(produtos, filters, usuario):
    """
    Salva o cálculo MRP na tabela do banco de dados.
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return False

    try:
        cursor = connection.cursor()

        for produto in produtos:
            if produto.get('op', 0) > 0:  # Salvar apenas produtos com OP preenchido
                # Truncar filter_value para evitar erro de varchar(255)
                filter_value = str(filters)
                if len(filter_value) > 250:
                    filter_value = filter_value[:247] + '...'

                cursor.execute("""
                    INSERT INTO mrp_calculation (
                        calculation_date, product_code, product_description, group_abbrev,
                        min_stock, economic_lot, current_stock, production_orders,
                        reservations, no_need, need, urgent, op_quantity,
                        balance_minus_reservation, future_balance, category,
                        filter_type, filter_value, created_by
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                """, (
                    datetime.now(),
                    produto['codigo'],
                    produto['descricao'],
                    produto['grupo'],
                    produto['estoque_minimo'],
                    produto['lote_economico'],
                    produto['estoque'],
                    produto['ordem'],
                    produto['reserva'],
                    produto['sem_necessidade'],
                    produto['necessidade'],
                    produto['urgente'],
                    produto['op'],
                    produto['saldo_menos_reserva'],
                    produto['saldo_futuro'],
                    produto['categoria'],
                    ','.join(filters.keys()),
                    filter_value,
                    usuario
                ))

        connection.commit()
        cursor.close()
        connection.close()
        return True

    except Error as e:
        print(f"Erro ao salvar MRP no banco: {e}")
        if connection:
            connection.rollback()
            connection.close()
        return False


def sanitize_excel_value(value):
    """
    Remove ou substitui caracteres inválidos para células do Excel.
    Excel não aceita certos caracteres de controle e símbolos especiais.
    """
    if value is None:
        return ''

    if not isinstance(value, str):
        return value

    # Remover caracteres de controle e outros problemáticos
    # Excel não aceita caracteres ASCII < 32 (exceto tab, newline, carriage return)
    sanitized = ''
    for char in value:
        char_code = ord(char)
        # Aceitar apenas caracteres válidos
        if char_code >= 32 or char in ['\t', '\n', '\r']:
            sanitized += char

    return sanitized


def generate_mrp_excel(produtos):
    """
    Gera arquivo Excel com os produtos que têm OP preenchido.

    Formato:
    - Coluna A: Código do Produto
    - Coluna B: Quantidade
    - Ordenado por categoria: Componentes, Assento/Encosto, Montado, Tampos
    """
    # Filtrar apenas produtos com OP > 0
    produtos_com_op = [p for p in produtos if p.get('op', 0) > 0]

    # Ordenar por categoria
    ordem_categorias = {
        'Componentes': 1,
        'Assento/Encosto': 2,
        'Montado': 3,
        'Tampos': 4
    }
    produtos_com_op.sort(key=lambda x: ordem_categorias.get(x.get('categoria', 'Componentes'), 5))

    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MRP"

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4F83CC", end_color="4F83CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cabeçalho (linha 1)
    ws['A1'] = 'Código'
    ws['B1'] = 'Quantidade'

    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = header_alignment

    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].alignment = header_alignment

    # Dados a partir da linha 2
    row = 2
    for produto in produtos_com_op:
        ws[f'A{row}'] = sanitize_excel_value(produto['codigo'])
        ws[f'B{row}'] = produto['op']
        row += 1

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    # Salvar em BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file


def init_mrp_tables():
    """
    Cria as tabelas necessárias para o sistema de sessões MRP no banco siga_db.
    Tabelas:
        - mrp_sessoes: Armazena as sessões de cálculo MRP
        - mrp_calculo_itens: Armazena os itens calculados de cada sessão
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        print("Erro: Não foi possível conectar ao banco siga_db para criar tabelas MRP")
        return False

    try:
        cursor = connection.cursor()

        # Criar tabela de sessões MRP
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS mrp_sessoes (
                id SERIAL PRIMARY KEY,
                descricao VARCHAR(500) NOT NULL,
                usuario VARCHAR(100) NOT NULL,
                data_criacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                data_calculo TIMESTAMP,
                data_atualizacao TIMESTAMP,
                status VARCHAR(20) DEFAULT 'pendente',
                filtros_json TEXT,
                observacoes TEXT
            )
        """)

        # Criar tabela de itens do cálculo MRP
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS mrp_calculo_itens (
                id SERIAL PRIMARY KEY,
                sessao_id INTEGER NOT NULL REFERENCES mrp_sessoes(id) ON DELETE CASCADE,
                codigo_produto VARCHAR(50) NOT NULL,
                descricao_produto VARCHAR(500),
                grupo VARCHAR(20),
                estoque_minimo DECIMAL(15,4) DEFAULT 0,
                lote_economico DECIMAL(15,4) DEFAULT 0,
                estoque DECIMAL(15,4) DEFAULT 0,
                ordem DECIMAL(15,4) DEFAULT 0,
                reserva DECIMAL(15,4) DEFAULT 0,
                sem_necessidade DECIMAL(15,4) DEFAULT 0,
                necessidade DECIMAL(15,4) DEFAULT 0,
                urgente DECIMAL(15,4) DEFAULT 0,
                op DECIMAL(15,4) DEFAULT 0,
                saldo_menos_reserva DECIMAL(15,4) DEFAULT 0,
                saldo_futuro DECIMAL(15,4) DEFAULT 0,
                categoria VARCHAR(50),
                material_type VARCHAR(50),
                cubagem_unitaria DECIMAL(15,6) DEFAULT 0,
                data_inclusao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                data_atualizacao TIMESTAMP
            )
        """)

        # Criar índices para melhor performance
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_mrp_sessoes_usuario ON mrp_sessoes(usuario)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_mrp_sessoes_data ON mrp_sessoes(data_criacao DESC)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_mrp_calculo_itens_sessao ON mrp_calculo_itens(sessao_id)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_mrp_calculo_itens_produto ON mrp_calculo_itens(codigo_produto)
        """)

        connection.commit()
        cursor.close()
        connection.close()
        print("Tabelas MRP criadas/verificadas com sucesso no banco siga_db")
        return True

    except Error as e:
        print(f"Erro ao criar tabelas MRP: {e}")
        if connection:
            connection.rollback()
            connection.close()
        return False


def criar_sessao_mrp(descricao, usuario, filtros=None):
    """
    Cria uma nova sessão de cálculo MRP com filtros pré-selecionados.

    Args:
        descricao: Descrição da sessão (ex: "OP 001 - PETRA")
        usuario: Nome do usuário que está criando
        filtros: Dict com filtros selecionados (pedidos, lotes_carga, lotes_producao)

    Returns:
        ID da sessão criada ou None em caso de erro
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return None

    try:
        import json
        filtros_json = json.dumps(filtros) if filtros else None

        cursor = connection.cursor()
        cursor.execute("""
            INSERT INTO mrp_sessoes (descricao, usuario, status, filtros_json)
            VALUES (%s, %s, 'pendente', %s)
            RETURNING id
        """, (descricao, usuario, filtros_json))

        sessao_id = cursor.fetchone()[0]
        connection.commit()
        cursor.close()
        connection.close()
        return sessao_id

    except Error as e:
        print(f"Erro ao criar sessão MRP: {e}")
        if connection:
            connection.rollback()
            connection.close()
        return None


def atualizar_status_sessao(sessao_id, novo_status):
    """
    Atualiza o status de uma sessão MRP.

    Args:
        sessao_id: ID da sessão
        novo_status: Novo status (pendente, calculado, editado, exportado)

    Returns:
        True se sucesso, False em caso de erro
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return False

    try:
        cursor = connection.cursor()
        cursor.execute("""
            UPDATE mrp_sessoes
            SET status = %s,
                data_atualizacao = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (novo_status, sessao_id))

        connection.commit()
        cursor.close()
        connection.close()
        return True

    except Error as e:
        print(f"Erro ao atualizar status da sessão: {e}")
        if connection:
            connection.rollback()
            connection.close()
        return False


def listar_sessoes_mrp(usuario=None, limite=50):
    """
    Lista as sessões de cálculo MRP.

    Args:
        usuario: Filtrar por usuário (opcional)
        limite: Quantidade máxima de sessões a retornar

    Returns:
        Lista de sessões
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return []

    try:
        cursor = connection.cursor()

        if usuario:
            cursor.execute("""
                SELECT id, descricao, usuario, data_criacao, data_calculo,
                       data_atualizacao, status, observacoes
                FROM mrp_sessoes
                WHERE usuario = %s
                ORDER BY data_criacao DESC
                LIMIT %s
            """, (usuario, limite))
        else:
            cursor.execute("""
                SELECT id, descricao, usuario, data_criacao, data_calculo,
                       data_atualizacao, status, observacoes
                FROM mrp_sessoes
                ORDER BY data_criacao DESC
                LIMIT %s
            """, (limite,))

        sessoes = []
        for row in cursor.fetchall():
            sessoes.append({
                'id': row[0],
                'descricao': row[1],
                'usuario': row[2],
                'data_criacao': row[3],
                'data_calculo': row[4],
                'data_atualizacao': row[5],
                'status': row[6],
                'observacoes': row[7]
            })

        cursor.close()
        connection.close()
        return sessoes

    except Error as e:
        print(f"Erro ao listar sessões MRP: {e}")
        if connection:
            connection.close()
        return []


def obter_sessao_mrp(sessao_id):
    """
    Obtém os detalhes de uma sessão MRP específica.

    Args:
        sessao_id: ID da sessão

    Returns:
        Dict com dados da sessão ou None
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return None

    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT id, descricao, usuario, data_criacao, data_calculo,
                   data_atualizacao, status, filtros_json, observacoes
            FROM mrp_sessoes
            WHERE id = %s
        """, (sessao_id,))

        row = cursor.fetchone()
        if not row:
            cursor.close()
            connection.close()
            return None

        sessao = {
            'id': row[0],
            'descricao': row[1],
            'usuario': row[2],
            'data_criacao': row[3],
            'data_calculo': row[4],
            'data_atualizacao': row[5],
            'status': row[6],
            'filtros_json': row[7],
            'observacoes': row[8]
        }

        cursor.close()
        connection.close()
        return sessao

    except Error as e:
        print(f"Erro ao obter sessão MRP: {e}")
        if connection:
            connection.close()
        return None


def salvar_calculo_mrp_sessao(sessao_id, produtos, filtros=None):
    """
    Salva os itens calculados do MRP em uma sessão.

    Args:
        sessao_id: ID da sessão
        produtos: Lista de produtos calculados
        filtros: Filtros utilizados no cálculo (opcional)

    Returns:
        True se sucesso, False em caso de erro
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return False

    try:
        cursor = connection.cursor()

        # Limpar itens anteriores da sessão (se existirem)
        cursor.execute("DELETE FROM mrp_calculo_itens WHERE sessao_id = %s", (sessao_id,))

        # Inserir novos itens
        for produto in produtos:
            cursor.execute("""
                INSERT INTO mrp_calculo_itens (
                    sessao_id, codigo_produto, descricao_produto, grupo,
                    estoque_minimo, lote_economico, estoque, ordem, reserva,
                    sem_necessidade, necessidade, urgente, op,
                    saldo_menos_reserva, saldo_futuro, categoria,
                    material_type, cubagem_unitaria
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """, (
                sessao_id,
                produto.get('codigo', ''),
                produto.get('descricao', ''),
                produto.get('grupo', ''),
                produto.get('estoque_minimo', 0),
                produto.get('lote_economico', 0),
                produto.get('estoque', 0),
                produto.get('ordem', 0),
                produto.get('reserva', 0),
                produto.get('sem_necessidade', 0),
                produto.get('necessidade', 0),
                produto.get('urgente', 0),
                produto.get('op', 0),
                produto.get('saldo_menos_reserva', 0),
                produto.get('saldo_futuro', 0),
                produto.get('categoria', ''),
                produto.get('material_type', ''),
                produto.get('cubagem_unitaria', 0)
            ))

        # Atualizar sessão com data do cálculo e filtros
        import json
        filtros_json = json.dumps(filtros) if filtros else None
        cursor.execute("""
            UPDATE mrp_sessoes
            SET data_calculo = CURRENT_TIMESTAMP,
                status = 'calculado',
                filtros_json = %s
            WHERE id = %s
        """, (filtros_json, sessao_id))

        connection.commit()
        cursor.close()
        connection.close()
        return True

    except Error as e:
        print(f"Erro ao salvar cálculo MRP na sessão: {e}")
        import traceback
        traceback.print_exc()
        if connection:
            connection.rollback()
            connection.close()
        return False


def atualizar_op_sessao(sessao_id, itens_op):
    """
    Atualiza os valores de OP dos itens de uma sessão.

    Args:
        sessao_id: ID da sessão
        itens_op: Dict {codigo_produto: valor_op}

    Returns:
        True se sucesso, False em caso de erro
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return False

    try:
        cursor = connection.cursor()

        for codigo, op in itens_op.items():
            # Calcular saldo_futuro = estoque + ordem - reserva + op
            cursor.execute("""
                UPDATE mrp_calculo_itens
                SET op = %s,
                    saldo_futuro = estoque + ordem - reserva + %s,
                    data_atualizacao = CURRENT_TIMESTAMP
                WHERE sessao_id = %s AND codigo_produto = %s
            """, (op, op, sessao_id, codigo))

        # Atualizar data de atualização da sessão
        cursor.execute("""
            UPDATE mrp_sessoes
            SET data_atualizacao = CURRENT_TIMESTAMP,
                status = 'editado'
            WHERE id = %s
        """, (sessao_id,))

        connection.commit()
        cursor.close()
        connection.close()
        return True

    except Error as e:
        print(f"Erro ao atualizar OP da sessão: {e}")
        if connection:
            connection.rollback()
            connection.close()
        return False


def obter_itens_sessao_mrp(sessao_id):
    """
    Obtém os itens calculados de uma sessão MRP.

    Args:
        sessao_id: ID da sessão

    Returns:
        Lista de produtos/itens
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return []

    try:
        cursor = connection.cursor()
        cursor.execute("""
            SELECT codigo_produto, descricao_produto, grupo, estoque_minimo,
                   lote_economico, estoque, ordem, reserva, sem_necessidade,
                   necessidade, urgente, op, saldo_menos_reserva, saldo_futuro,
                   categoria, material_type, cubagem_unitaria
            FROM mrp_calculo_itens
            WHERE sessao_id = %s
            ORDER BY grupo, codigo_produto
        """, (sessao_id,))

        produtos = []
        for row in cursor.fetchall():
            produtos.append({
                'codigo': row[0],
                'descricao': row[1],
                'grupo': row[2],
                'estoque_minimo': float(row[3] or 0),
                'lote_economico': float(row[4] or 0),
                'estoque': float(row[5] or 0),
                'ordem': float(row[6] or 0),
                'reserva': float(row[7] or 0),
                'sem_necessidade': float(row[8] or 0),
                'necessidade': float(row[9] or 0),
                'urgente': float(row[10] or 0),
                'op': float(row[11] or 0),
                'saldo_menos_reserva': float(row[12] or 0),
                'saldo_futuro': float(row[13] or 0),
                'categoria': row[14],
                'material_type': row[15],
                'cubagem_unitaria': float(row[16] or 0)
            })

        cursor.close()
        connection.close()
        return produtos

    except Error as e:
        print(f"Erro ao obter itens da sessão MRP: {e}")
        if connection:
            connection.close()
        return []


def excluir_sessao_mrp(sessao_id):
    """
    Exclui uma sessão MRP e seus itens.

    Args:
        sessao_id: ID da sessão

    Returns:
        True se sucesso, False em caso de erro
    """
    connection = get_db_connection(SIGA_DB_NAME)
    if not connection:
        return False

    try:
        cursor = connection.cursor()

        # Os itens serão excluídos automaticamente por ON DELETE CASCADE
        cursor.execute("DELETE FROM mrp_sessoes WHERE id = %s", (sessao_id,))

        connection.commit()
        cursor.close()
        connection.close()
        return True

    except Error as e:
        print(f"Erro ao excluir sessão MRP: {e}")
        if connection:
            connection.rollback()
            connection.close()
        return False


def generate_mrp_excel_all(produtos):
    """
    Gera arquivo Excel com TODOS os produtos calculados.

    Formato completo com todas as colunas do cálculo MRP:
    - Grupo, Código, Descrição, Est. Mín., Lote Eco., Estoque, Ordem, Reserva,
      Sem Necessidade, Necessidade, Urgente, OP, Saldo - Res., Saldo Futuro
    """
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MRP Completo"

    # Estilo do cabeçalho padrão
    header_fill = PatternFill(start_color="4F83CC", end_color="4F83CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Estilos especiais para colunas coloridas
    estoque_header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    ordem_header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    reserva_header_fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")

    estoque_body_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    ordem_body_fill = PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid")
    reserva_body_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    # Cabeçalhos
    headers = [
        'Grupo', 'Código', 'Descrição', 'Est. Mín.', 'Lote Eco.',
        'Estoque', 'Ordem', 'Reserva', 'Sem Necess.', 'Necessidade',
        'Urgente', 'OP', 'Saldo - Res.', 'Saldo Futuro'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment

        # Aplicar cor específica para Estoque, Ordem e Reserva
        if col_idx == 6:  # Estoque
            cell.fill = estoque_header_fill
        elif col_idx == 7:  # Ordem
            cell.fill = ordem_header_fill
        elif col_idx == 8:  # Reserva
            cell.fill = reserva_header_fill
        else:
            cell.fill = header_fill

    # Dados a partir da linha 2
    row = 2
    for produto in produtos:
        ws.cell(row=row, column=1).value = sanitize_excel_value(produto.get('grupo', ''))
        ws.cell(row=row, column=2).value = sanitize_excel_value(produto.get('codigo', ''))
        ws.cell(row=row, column=3).value = sanitize_excel_value(produto.get('descricao', ''))
        ws.cell(row=row, column=4).value = produto.get('estoque_minimo', 0)
        ws.cell(row=row, column=5).value = produto.get('lote_economico', 0)

        # Estoque com cor verde
        cell_estoque = ws.cell(row=row, column=6)
        cell_estoque.value = produto.get('estoque', 0)
        cell_estoque.fill = estoque_body_fill

        # Ordem com cor azul
        cell_ordem = ws.cell(row=row, column=7)
        cell_ordem.value = produto.get('ordem', 0)
        cell_ordem.fill = ordem_body_fill

        # Reserva com cor amarela
        cell_reserva = ws.cell(row=row, column=8)
        cell_reserva.value = produto.get('reserva', 0)
        cell_reserva.fill = reserva_body_fill

        ws.cell(row=row, column=9).value = produto.get('sem_necessidade', 0)
        ws.cell(row=row, column=10).value = produto.get('necessidade', 0)
        ws.cell(row=row, column=11).value = produto.get('urgente', 0)
        ws.cell(row=row, column=12).value = produto.get('op', 0)

        # Saldo - Reserva com fórmula: Estoque + Ordem - Reserva (F + G - H)
        cell_saldo_reserva = ws.cell(row=row, column=13)
        cell_saldo_reserva.value = f'=F{row}+G{row}-H{row}'

        # Saldo Futuro com fórmula: Estoque + Ordem - Reserva + OP (F + G - H + L)
        cell_saldo_futuro = ws.cell(row=row, column=14)
        cell_saldo_futuro.value = f'=F{row}+G{row}-H{row}+L{row}'

        row += 1

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 8   # Grupo
    ws.column_dimensions['B'].width = 12  # Código
    ws.column_dimensions['C'].width = 35  # Descrição
    ws.column_dimensions['D'].width = 10  # Est. Mín.
    ws.column_dimensions['E'].width = 10  # Lote Eco.
    ws.column_dimensions['F'].width = 10  # Estoque
    ws.column_dimensions['G'].width = 10  # Ordem
    ws.column_dimensions['H'].width = 10  # Reserva
    ws.column_dimensions['I'].width = 12  # Sem Necess.
    ws.column_dimensions['J'].width = 12  # Necessidade
    ws.column_dimensions['K'].width = 10  # Urgente
    ws.column_dimensions['L'].width = 10  # OP
    ws.column_dimensions['M'].width = 12  # Saldo - Res.
    ws.column_dimensions['N'].width = 12  # Saldo Futuro

    # Salvar em BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file
